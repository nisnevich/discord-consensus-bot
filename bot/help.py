import logging
import csv
import io
import discord
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font, Border, Side, PatternFill, Color
from openpyxl.styles.alignment import Alignment
from sqlalchemy import and_

from bot.config.const import *
from bot.config.logging_config import log_handler, console_handler
from bot.utils.discord_utils import get_discord_client, get_message, get_user_by_id_or_mention
from bot.utils.validation import validate_roles
from bot.utils.db_utils import DBUtil
from bot.utils.formatting_utils import get_amount_to_print, get_nickname_by_id_or_mention
from bot.config.schemas import (
    Proposals,
    ProposalHistory,
    FreeFundingTransaction,
    FreeFundingBalance,
)
from bot.config.const import ProposalResult, VOTING_CHANNEL_ID

logger = logging.getLogger(__name__)
logger.setLevel(DEFAULT_LOG_LEVEL)
logger.addHandler(log_handler)
logger.addHandler(console_handler)

db = DBUtil()
client = get_discord_client()

# Create alignments to format cells
alignment_center = Alignment(horizontal='center', vertical='center')
alignment_wrap = Alignment(wrap_text=True)
alignment_wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
alignment_left_center = Alignment(horizontal='left', vertical='center')
# Create borders
bottom_border = Border(bottom=Side(style='thin'))
dotted_right_border = Border(right=Side(style='dotted'))
# Create colors
header_fill = PatternFill(start_color=Color('b6d7a8'), end_color=Color('b6d7a8'), fill_type='solid')


@client.event
async def on_message(message):
    # If the author is not bot, and the message is written in the voting channel, remove the message and DM user, explaining why
    # This is to maintain the channel cleaner
    if (
        REMOVE_HUMAN_MESSAGES_FROM_VOTING_CHANNEL
        and message.channel.id == VOTING_CHANNEL_ID
        and not message.author.bot
    ):
        await message.author.send(HELP_MESSAGE_REMOVED_FROM_VOTING_CHANNEL)
        await message.delete()
        return

    # Replace curly quotes with standard quotes - otherwise it causes errors when used like !gift @Consensus-Dev#0594 500 don’t worry, be happy :blush:
    message.content = message.content.replace('’', "'")
    # Process commands (the @client.event decorator intercepts all messages)
    await client.process_commands(message)


@client.command(name=FREE_FUNDING_BALANCE_COMMAND_NAME, aliases=FREE_FUNDING_BALANCE_ALIASES)
async def send_free_funding_balance(ctx):
    try:
        # Reply to a non-authorized user
        if not await validate_roles(ctx.message.author):
            # Adding greetings and "cancelled" reactions
            await ctx.message.add_reaction(REACTION_ON_BOT_MENTION)
            # Sending response in DM
            await ctx.author.send(HELP_MESSAGE_NON_AUTHORIZED_USER)
            return

        # Retrieve the authors balance
        author_balance = db.get_user_free_funding_balance(ctx.message.author.id)
        # Add author to DB if not added yet
        if not author_balance:
            logger.debug("Added free funding balance for author=%s", ctx.message.author.id)
            author_balance = FreeFundingBalance(
                author_id=ctx.message.author.id,
                author_nickname=await get_nickname_by_id_or_mention(
                    str(ctx.message.author.mention)
                ),
                balance=FREE_FUNDING_LIMIT_PERSON_PER_SEASON,
            )
            await db.add(author_balance)

        # Reply with the balance
        await ctx.message.add_reaction(REACTION_ON_BOT_MENTION)
        await ctx.message.reply(
            FREE_FUNDING_BALANCE_MESSAGE.format(balance=get_amount_to_print(author_balance.balance))
        )

    except Exception as e:
        try:
            # Try replying in Discord
            await ctx.message.reply(
                f"An unexpected error occurred when sending free funding balance. cc {RESPONSIBLE_MENTION}"
            )
        except Exception as e:
            logger.critical("Unable to reply in the chat that a critical error has occurred.")

        logger.critical(
            "Unexpected error in %s while sending help, channel=%s, message=%s, user=%s",
            __name__,
            ctx.message.channel.id if ctx.message.channel else None,
            ctx.message.id,
            ctx.message.author.mention,
            exc_info=True,
        )


@client.command(name=HELP_COMMAND_NAME, aliases=HELP_COMMAND_ALIASES)
async def help(ctx):
    try:
        # Remove the help request message
        await ctx.message.delete()
        # Reply to a non-authorized user
        if not await validate_roles(ctx.message.author):
            await ctx.author.send(HELP_MESSAGE_NON_AUTHORIZED_USER)
            return
        # Reply to an authorized user
        message = await ctx.author.send(HELP_MESSAGE_AUTHORIZED_USER)
        await message.edit(suppress=True)
    except Exception as e:
        try:
            # Try replying in Discord
            await ctx.message.reply(
                f"An unexpected error occurred when sending help. cc {RESPONSIBLE_MENTION}"
            )
        except Exception as e:
            logger.critical("Unable to reply in the chat that a critical error has occurred.")

        logger.critical(
            "Unexpected error in %s while sending help, channel=%s, message=%s, user=%s",
            __name__,
            ctx.message.channel.id if ctx.message.channel else None,
            ctx.message.id,
            ctx.message.author.mention,
            exc_info=True,
        )


def define_columns(page, columns):
    """
    Columns are defined the same way for each page, with bold headers.
    """
    # Write the column names to the worksheet and set column widths
    for col_num, column in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        column_header = column["header"]
        column_width = column["width"]
        page.column_dimensions[column_letter].width = column_width
        page.cell(row=1, column=col_num, value=column_header).font = Font(bold=True)

    # Draw the border after the header
    for column in range(1, 1 + len(columns)):
        cell = page.cell(row=1, column=column)
        cell.border = bottom_border
        cell.fill = header_fill

    # Freeze the header to show it when scrolling the sheet (it freezes all rows above the given cell)
    page.freeze_panes = 'A2'


def postprocess(page, columns):
    """
    Format page after it has been filled with the data.
    """
    row_num = 0
    # Apply a dotted right border style to all cells in the specified worksheet, while preserving any existing bottom borders.
    for row in page.iter_rows():
        row_num += 1
        # Skip the first row, as it's the header
        if row_num == 1:
            continue
        for cell in row:
            if cell.border.bottom and cell.border.bottom.style:
                cell.border = Border(bottom=cell.border.bottom, right=dotted_right_border.right)
            else:
                cell.border = dotted_right_border


def set_bottom_border(page, columns):
    """
    Applies a bottom border to all cells in the last row of the specified worksheet for each column in the specified columns list.
    """
    # Loop over the columns and apply the border to each cell
    for column in range(1, 1 + len(columns)):
        cell = page.cell(row=page.max_row, column=column)
        cell.border = bottom_border


async def write_lazy_consensus_history(page):
    # Define column names and widths
    columns = [
        {"header": "Discord link", "width": 15},
        {"header": "When completed (UTC time)", "width": 28},
        {"header": "Author", "width": 20},
        {"header": "Grant given to", "width": 25},
        {"header": "Amount", "width": 10},
        {"header": "Total amount", "width": 14},
        # Description isn't aligned at full height of the proposal text. Unfortunately, in openpyxl
        # there seems to be no way to do so, using column_dimensions.auto_height didn't help. The
        # workaround could be to adjust column width based on the font size and number of lines.
        {"header": "Description", "width": 100},
    ]
    # Enable the columns in the page
    define_columns(page, columns)

    # Retrieve all accepted proposals
    accepted_proposals = await db.filter(
        ProposalHistory,
        condition=and_(
            Proposals.id == ProposalHistory.id,
            ProposalHistory.result == ProposalResult.ACCEPTED.value,
        ),
        order_by=ProposalHistory.closed_at.asc(),
    )
    # Loop over each accepted proposal and add a row to the worksheet
    current_row = 2
    for proposal in accepted_proposals.all():
        logger.debug(proposal)
        start_row = end_row = current_row
        # If the proposal is not financial, fill recievers and amount with empty analytics values
        if proposal.not_financial:
            cell = page.cell(row=start_row, column=4, value=EMPTY_ANALYTICS_VALUE)
            cell.alignment = alignment_center
            cell = page.cell(row=start_row, column=5, value=EMPTY_ANALYTICS_VALUE)
            cell.alignment = alignment_center
        else:
            # Retrieve recievers
            finance_recipients = proposal.finance_recipients
            if not finance_recipients:
                logger.warning("No finance recipients found in a financial proposal!")
                continue
            # Fill each receivers group in a separate row
            for row_num, recipient in enumerate(finance_recipients, start_row):
                # Mention
                cell = page.cell(
                    row=row_num,
                    column=4,
                    value=str(recipient.recipient_nicknames),
                )
                cell.alignment = alignment_wrap
                # Amount
                cell = page.cell(
                    row=row_num, column=5, value=str(get_amount_to_print(recipient.amount))
                )
                cell.alignment = alignment_center
            # Set the end row of the current proposal to merge other cells
            end_row += len(finance_recipients) - 1

        # Discord URL
        discord_link = proposal.voting_message_url
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
        cell = page.cell(row=start_row, column=1, value=discord_link)
        cell.hyperlink = discord_link
        cell.alignment = alignment_left_center
        # Date
        date_str = proposal.closed_at.strftime("%Y-%m-%d %H:%M:%S")
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
        cell = page.cell(row=start_row, column=2, value=date_str)
        cell.alignment = alignment_center
        # Author
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=3, end_column=3)
        cell = page.cell(
            row=start_row,
            column=3,
            value=str(proposal.author_nickname),
        )
        cell.alignment = alignment_wrap_center
        # Total amount
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=6, end_column=6)
        cell = page.cell(
            row=start_row,
            column=6,
            value=EMPTY_ANALYTICS_VALUE
            if proposal.not_financial
            else get_amount_to_print(proposal.total_amount),
        )
        cell.alignment = alignment_center
        # Description
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=7, end_column=7)
        cell = page.cell(row=start_row, column=7, value=str(proposal.description))
        cell.alignment = alignment_wrap

        # Draw the bottom border
        set_bottom_border(page, columns)

        # Increment the current row
        current_row = end_row + 1

    # Apply some formatting after the page has been filled
    postprocess(page, columns)


async def write_free_funding_balance(page):
    # Define column names and widths
    columns = [
        {"header": "Author", "width": 15},
        {"header": "Remaining balance", "width": 20},
    ]
    # Enable the columns in the page
    define_columns(page, columns)
    # Write a note in the corner
    page.cell(
        row=1, column=3, value="Note: only users that have used tips in this season are listed here"
    ).font = Font(italic=True)

    # Retrieve all balances
    all_balances = await db.filter(FreeFundingBalance, is_history=False)

    # Loop over each users balance and add a row to the worksheet
    for row_num, balance in enumerate(all_balances.all(), 2):
        # Author
        page.cell(row=row_num, column=1, value=str(balance.author_nickname))
        # Amount
        page.cell(row=row_num, column=2, value=str(get_amount_to_print(balance.balance)))

        # Draw the bottom border
        set_bottom_border(page, columns)

    # Apply some formatting after the page has been filled
    postprocess(page, columns)


async def write_free_funding_transactions(page):
    # Define column names and widths
    columns = [
        {"header": "Discord link", "width": 15},
        {"header": "UTC time", "width": 20},
        {"header": "Author", "width": 15},
        {"header": "Sent to", "width": 20},
        {"header": "Total amount", "width": 15},
        {"header": "Description", "width": 100},
    ]
    # Enable the columns in the page
    define_columns(page, columns)

    # Retrieve all transactions
    all_transactions = await db.filter(
        FreeFundingTransaction,
        order_by=FreeFundingTransaction.submitted_at.asc(),
    )

    # Loop over each transaction and add a row to the worksheet
    for row_num, transaction in enumerate(all_transactions.all(), 2):
        # Discord URL
        discord_link = transaction.message_url
        page.cell(row=row_num, column=1).value = discord_link
        page.cell(row=row_num, column=1).hyperlink = discord_link
        # Date
        page.cell(
            row=row_num, column=2, value=transaction.submitted_at.strftime("%Y-%m-%d %H:%M:%S")
        )
        # Author
        page.cell(row=row_num, column=3, value=str(transaction.author_nickname))
        # Mentions
        page.cell(row=row_num, column=4, value=str(transaction.recipient_nicknames))
        # Total amount
        page.cell(row=row_num, column=5, value=str(get_amount_to_print(transaction.total_amount)))
        # Description
        page.cell(row=row_num, column=6, value=str(transaction.description))

        # Draw the bottom border
        set_bottom_border(page, columns)

    # Apply some formatting after the page has been filled
    postprocess(page, columns)


async def export_xlsx():
    # Create a new Excel workbook and worksheet
    wb = openpyxl.Workbook()

    # Create a page with a history of all lazy consensus proposals
    lazy_history_page = wb.active
    lazy_history_page.title = "Lazy Consensus History"
    await write_lazy_consensus_history(lazy_history_page)

    # Create a page with free funding balances of all members
    free_funding_balance_page = wb.create_sheet(title="Tips Balances")
    await write_free_funding_balance(free_funding_balance_page)

    # Create a page with all free funding transactions
    free_funding_history_page = wb.create_sheet(title="Tips Transactions")
    await write_free_funding_transactions(free_funding_history_page)

    # Save the Excel workbook to a temporary file
    temp_file = io.BytesIO()
    wb.save(temp_file)
    temp_file.seek(0)

    return temp_file, EXPORT_DATA_FILENAME


@client.command(name=EXPORT_COMMAND_NAME)
async def export(ctx):
    try:
        # Reply to a non-authorized user
        if not await validate_roles(ctx.message.author):
            # Adding greetings and "cancelled" reactions
            await ctx.message.add_reaction(REACTION_ON_BOT_MENTION)
            # Sending response in DM
            await ctx.message.reply(HELP_MESSAGE_NON_AUTHORIZED_USER)
            return
        # Adding greetings reaction so to show that the command is being processed (it may take a couple of seconds waiting for the user)
        await ctx.message.add_reaction(REACTION_ON_BOT_MENTION)

        # Create the document
        document, filename = await export_xlsx()
        # Send the document to user
        await ctx.message.reply(
            EXPORT_CHANNEL_REPLY,
            file=discord.File(document, filename=filename),
        )

    except Exception as e:
        try:
            # Try replying in Discord
            await ctx.message.reply(
                f"An unexpected error occurred when exporting analytical data. cc {RESPONSIBLE_MENTION}"
            )
        except Exception as e:
            logger.critical("Unable to reply in the chat that a critical error has occurred.")

        logger.critical(
            "Unexpected error in %s while exporting analytical data, channel=%s, message=%s, user=%s",
            __name__,
            ctx.message.channel.id if ctx.message.channel else None,
            ctx.message.id,
            ctx.message.author.mention,
            exc_info=True,
        )
