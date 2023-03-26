import logging
import io
import discord
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font, Border, Side, PatternFill, Color
from openpyxl.styles.alignment import Alignment
from openpyxl.chart import LineChart, Reference, Series

from sqlalchemy import and_
from datetime import datetime

from bot.config.const import *
from bot.config.logging_config import log_handler, console_handler
from bot.utils.discord_utils import get_discord_client, get_message, get_user_by_id_or_mention
from bot.utils.validation import validate_roles
from bot.utils.db_utils import DBUtil
from bot.utils.formatting_utils import get_amount_to_print, get_nickname_by_id_or_mention
from bot.config.schemas import (
    Proposals,
    ProposalHistory,
    FreeFundingTransaction,
    FreeFundingBalance,
    FinanceRecipients,
    Voters,
)
from bot.config.const import ProposalResult, VOTING_CHANNEL_ID

logger = logging.getLogger(__name__)
logger.setLevel(DEFAULT_LOG_LEVEL)
logger.addHandler(log_handler)
logger.addHandler(console_handler)

db = DBUtil()
client = get_discord_client()

# Create alignments to format cells
alignment_center = Alignment(horizontal='center', vertical='center')
alignment_wrap = Alignment(wrap_text=True)
alignment_wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
alignment_left_center = Alignment(horizontal='left', vertical='center')
# Create borders
bottom_border = Border(bottom=Side(style='thin'))
dotted_right_border = Border(right=Side(style='dotted'))
# Create colors
header_fill = PatternFill(start_color=Color('b6d7a8'), end_color=Color('b6d7a8'), fill_type='solid')
header_font = Font(bold=True)


def define_columns(page, columns):
    """
    Columns are defined the same way for each page, with bold headers.
    """
    # Write the column names to the worksheet and set column widths
    for col_num, column in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        column_header = column["header"]
        column_width = column["width"]
        page.column_dimensions[column_letter].width = column_width
        page.cell(row=1, column=col_num, value=column_header).font = header_font

    # Draw the border after the header
    for column in range(1, 1 + len(columns)):
        cell = page.cell(row=1, column=column)
        cell.border = bottom_border
        cell.fill = header_fill

    # Freeze the header to show it when scrolling the sheet (it freezes all rows above the given cell)
    page.freeze_panes = 'A2'


def postprocess(page):
    """
    Format page after it has been filled with the data.
    """
    row_num = 0
    # Apply a dotted right border style to all cells in the specified worksheet, while preserving any existing bottom borders.
    for row in page.iter_rows():
        row_num += 1
        # Skip the first row, as it's the header
        if row_num == 1:
            continue
        for cell in row:
            if cell.border.bottom and cell.border.bottom.style:
                cell.border = Border(bottom=cell.border.bottom, right=dotted_right_border.right)
            else:
                cell.border = dotted_right_border


def set_bottom_border(page, columns):
    """
    Applies a bottom border to all cells in the last row of the specified worksheet for each column in the specified columns list.
    """
    # Loop over the columns and apply the border to each cell
    for column in range(1, 1 + (columns if isinstance(columns, int) else len(columns))):
        cell = page.cell(row=page.max_row, column=column)
        cell.border = bottom_border


async def get_unique_active_users():
    """
    Returns a list of unique users who have used free funding, submitted proposals or voted.
    """
    users_with_tips = {
        (balance.author_id, balance.author_nickname)
        for balance in await db.filter(FreeFundingBalance, is_history=False)
    }
    users_with_proposals = {
        (proposal.author_id, proposal.author_nickname)
        for proposal in await db.filter(ProposalHistory)
    }
    users_with_votes = {(voter.user_id, voter.user_nickname) for voter in await db.filter(Voters)}
    return users_with_tips.union(users_with_proposals).union(users_with_votes)


async def write_summary(page):
    # Number of rows with data in summary page
    number_of_summary_rows = 4
    # Width of first column (fields desciptions)
    first_col_width = 28
    # Set first row width
    page.column_dimensions[get_column_letter(1)].width = first_col_width
    # Set first row formatting
    for row in range(1, number_of_summary_rows + 1):
        cell = page.cell(row=row, column=1)
        cell.alignment = alignment_wrap_center
        cell.font = header_font
        cell.fill = header_fill

    # Retrieve the necessary data
    free_funding_balances = await db.filter(FreeFundingBalance, is_history=False)
    free_funding_spent = sum(
        [
            FREE_FUNDING_LIMIT_PERSON_PER_SEASON - balance.balance
            for balance in free_funding_balances
        ]
    )
    accepted_proposals = await db.filter(
        ProposalHistory, condition=ProposalHistory.result == ProposalResult.ACCEPTED.value
    )
    total_grants_amount = sum(
        [proposal.total_amount for proposal in accepted_proposals if not proposal.not_financial]
    )
    total_accepted_proposals = accepted_proposals.count()

    submitted_proposals = await db.filter(Proposals)
    total_submitted_proposals = submitted_proposals.count()

    # Write the data to the page
    page.cell(row=1, column=1, value="Total tips sent:")
    page.cell(row=1, column=2, value=free_funding_spent)
    set_bottom_border(page, 2)

    page.cell(row=2, column=1, value="Total points given with full consensus:")
    page.cell(row=2, column=2, value=total_grants_amount)
    set_bottom_border(page, 2)

    page.cell(row=3, column=1, value="Total number of accepted proposals:")
    page.cell(row=3, column=2, value=total_accepted_proposals)
    set_bottom_border(page, 2)

    page.cell(row=4, column=1, value="Total number of submitted proposals:")
    page.cell(row=4, column=2, value=total_submitted_proposals)
    set_bottom_border(page, 2)


async def write_user_activity(page):
    """
    Creates a page with the activity of users who are allowed to send free funding, submit proposals
    and vote.
    """
    # Define column names and widths
    columns = [
        {"header": "Author", "width": 15},
        {"header": "Remaining tips", "width": 18},
        {"header": "Accepted proposals", "width": 20},
        {"header": "Submitted proposals", "width": 20},
        {"header": "Votes", "width": 15},
    ]
    # Enable the columns in the page
    define_columns(page, columns)

    # Retrieve all unique user IDs who used tips, submitted proposals, or voted and their nicknames
    unique_active_users = await get_unique_active_users()

    # Loop over each user and add a row to the worksheet
    for row_num, (user_id, user_nickname) in enumerate(
        sorted(unique_active_users, key=lambda x: x[0]), 2
    ):
        balances = await db.filter(FreeFundingBalance, is_history=False)
        # If the user haven't used free funding before, show his balance as default (we could have
        # added his balance to db here, but it's not the best place to do so in analytics)
        user_balance = FREE_FUNDING_LIMIT_PERSON_PER_SEASON
        for balance in balances.all():
            if balance.author_id == user_id:
                user_balance = balance.balance
                break

        # User
        page.cell(row=row_num, column=1, value=str(user_nickname))
        # Free funding balance
        page.cell(row=row_num, column=2, value=str(get_amount_to_print(user_balance)))
        # Accepted proposals
        accepted_proposals = await db.filter(
            ProposalHistory,
            condition=(ProposalHistory.author_id == user_id)
            & (ProposalHistory.result == ProposalResult.ACCEPTED.value),
        )
        page.cell(row=row_num, column=3, value=accepted_proposals.count())
        # Submitted proposals
        submitted_proposals = await db.filter(Proposals, condition=Proposals.author_id == user_id)
        page.cell(row=row_num, column=4, value=submitted_proposals.count())
        # Votes
        votes = await db.filter(Voters, condition=Voters.user_id == user_id)
        page.cell(row=row_num, column=5, value=votes.count())

    # Draw the bottom border
    set_bottom_border(page, columns)

    # Apply some formatting after the page has been filled
    postprocess(page)


async def write_user_grants_recieved(page):
    # Define column names and widths
    columns = [
        {"header": "User", "width": 20},
        #  {"header": "Points balance", "width": 20},
        {"header": "Tips received", "width": 14},
        {"header": "Grants received (by voting)", "width": 27},
    ]
    # Enable the columns in the page
    define_columns(page, columns)

    # Retrieve free funding transactions and group by user
    free_funding_transactions = await db.filter(FreeFundingTransaction)
    free_funding_by_user = {}
    for transaction in free_funding_transactions:
        recipients = transaction.recipient_nicknames.split(DB_ARRAY_COLUMN_SEPARATOR)
        amounts = [transaction.total_amount / len(recipients)] * len(recipients)
        for recipient, amount in zip(recipients, amounts):
            if recipient in free_funding_by_user:
                free_funding_by_user[recipient] += amount
            else:
                free_funding_by_user[recipient] = amount

    # Retrieve accepted proposals from ProposalHistory
    accepted_proposals = await db.filter(
        ProposalHistory, condition=ProposalHistory.result == ProposalResult.ACCEPTED.value
    )
    # Get the finance recipient IDs of the accepted proposals
    accepted_proposal_ids = [proposal.id for proposal in accepted_proposals]
    # Retrieve finance recipients whose proposal_id is in the list of accepted_proposal_ids
    finance_recipients = await db.filter(
        FinanceRecipients, condition=FinanceRecipients.proposal_id.in_(accepted_proposal_ids)
    )
    # Group by user
    grants_by_user = {}
    for recipient in finance_recipients:
        recipients = recipient.recipient_nicknames.split(COMMA_LIST_SEPARATOR)
        amounts = [recipient.amount] * len(recipients)
        for recipient, amount in zip(recipients, amounts):
            if recipient in grants_by_user:
                grants_by_user[recipient] += amount
            else:
                grants_by_user[recipient] = amount

    # Combine users from both dictionaries
    all_users = set(free_funding_by_user.keys()).union(grants_by_user.keys())

    # Write user data to the page
    row = 2
    for user in sorted(all_users):
        # Username
        page.cell(row=row, column=1, value=user)
        #
        page.cell(row=row, column=2, value=get_amount_to_print(free_funding_by_user.get(user, 0)))
        page.cell(row=row, column=3, value=get_amount_to_print(grants_by_user.get(user, 0)))
        # Draw the bottom border
        set_bottom_border(page, columns)

        row += 1

    # Apply some formatting after the page has been filled
    postprocess(page)


async def write_lazy_consensus_history(page):
    # Define column names and widths
    columns = [
        {"header": "Discord link", "width": 15},
        {"header": "When completed (UTC time)", "width": 28},
        {"header": "Author", "width": 20},
        {"header": "Grant given to", "width": 25},
        {"header": "Amount", "width": 10},
        {"header": "Total amount", "width": 14},
        # Description isn't aligned at full height of the proposal text. Unfortunately, in openpyxl
        # there seems to be no way to do so, using column_dimensions.auto_height didn't help. The
        # workaround could be to adjust column width based on the font size and number of lines.
        {"header": "Description", "width": 100},
    ]
    # Enable the columns in the page
    define_columns(page, columns)

    # Retrieve all accepted proposals
    accepted_proposals = await db.filter(
        ProposalHistory,
        condition=and_(
            Proposals.id == ProposalHistory.id,
            ProposalHistory.result == ProposalResult.ACCEPTED.value,
        ),
        order_by=ProposalHistory.closed_at.asc(),
    )
    # Loop over each accepted proposal and add a row to the worksheet
    current_row = 2
    for proposal in accepted_proposals.all():
        logger.debug(proposal)
        start_row = end_row = current_row
        # If the proposal is not financial, fill recievers and amount with empty analytics values
        if proposal.not_financial:
            cell = page.cell(row=start_row, column=4, value=EMPTY_ANALYTICS_VALUE)
            cell.alignment = alignment_center
            cell = page.cell(row=start_row, column=5, value=EMPTY_ANALYTICS_VALUE)
            cell.alignment = alignment_center
        else:
            # Retrieve recievers
            finance_recipients = proposal.finance_recipients
            if not finance_recipients:
                logger.warning("No finance recipients found in a financial proposal!")
                continue
            # Fill each receivers group in a separate row
            for row_num, recipient in enumerate(finance_recipients, start_row):
                # Mention
                cell = page.cell(
                    row=row_num,
                    column=4,
                    value=str(recipient.recipient_nicknames),
                )
                cell.alignment = alignment_wrap
                # Amount
                cell = page.cell(
                    row=row_num, column=5, value=str(get_amount_to_print(recipient.amount))
                )
                cell.alignment = alignment_center
            # Set the end row of the current proposal to merge other cells
            end_row += len(finance_recipients) - 1

        # Discord URL
        discord_link = proposal.voting_message_url
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
        cell = page.cell(row=start_row, column=1, value=discord_link)
        cell.hyperlink = discord_link
        cell.alignment = alignment_left_center
        # Date
        date_str = proposal.closed_at.strftime("%Y-%m-%d %H:%M:%S")
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
        cell = page.cell(row=start_row, column=2, value=date_str)
        cell.alignment = alignment_center
        # Author
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=3, end_column=3)
        cell = page.cell(
            row=start_row,
            column=3,
            value=str(proposal.author_nickname),
        )
        cell.alignment = alignment_wrap_center
        # Total amount
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=6, end_column=6)
        cell = page.cell(
            row=start_row,
            column=6,
            value=EMPTY_ANALYTICS_VALUE
            if proposal.not_financial
            else get_amount_to_print(proposal.total_amount),
        )
        cell.alignment = alignment_center
        # Description
        page.merge_cells(start_row=start_row, end_row=end_row, start_column=7, end_column=7)
        cell = page.cell(row=start_row, column=7, value=str(proposal.description))
        cell.alignment = alignment_wrap

        # Draw the bottom border
        set_bottom_border(page, columns)

        # Increment the current row
        current_row = end_row + 1

    # Apply some formatting after the page has been filled
    postprocess(page)


async def write_free_funding_transactions(page):
    # Define column names and widths
    columns = [
        {"header": "Discord link", "width": 15},
        {"header": "UTC time", "width": 20},
        {"header": "Author", "width": 15},
        {"header": "Sent to", "width": 20},
        {"header": "Total amount", "width": 15},
        {"header": "Description", "width": 100},
    ]
    # Enable the columns in the page
    define_columns(page, columns)

    # Retrieve all transactions
    all_transactions = await db.filter(
        FreeFundingTransaction,
        order_by=FreeFundingTransaction.submitted_at.asc(),
    )

    # Loop over each transaction and add a row to the worksheet
    for row_num, transaction in enumerate(all_transactions.all(), 2):
        # Discord URL
        discord_link = transaction.message_url
        page.cell(row=row_num, column=1).value = discord_link
        page.cell(row=row_num, column=1).hyperlink = discord_link
        # Date
        page.cell(
            row=row_num, column=2, value=transaction.submitted_at.strftime("%Y-%m-%d %H:%M:%S")
        )
        # Author
        page.cell(row=row_num, column=3, value=str(transaction.author_nickname))
        # Mentions
        page.cell(row=row_num, column=4, value=str(transaction.recipient_nicknames))
        # Total amount
        page.cell(row=row_num, column=5, value=str(get_amount_to_print(transaction.total_amount)))
        # Description
        page.cell(row=row_num, column=6, value=str(transaction.description))

        # Draw the bottom border
        set_bottom_border(page, columns)

    # Apply some formatting after the page has been filled
    postprocess(page)


async def export_xlsx():
    # Create a new Excel workbook and worksheet
    wb = openpyxl.Workbook()

    # Create a summary page
    summary_page = wb.active
    summary_page.title = "Summary"
    await write_summary(summary_page)

    # Create a page with free funding balances of all members
    free_funding_balance_page = wb.create_sheet(title="L3 Activity")
    await write_user_activity(free_funding_balance_page)

    # Create a page with grants and free funding received by all members
    all_grants_page = wb.create_sheet(title="Grant Receivers")
    await write_user_grants_recieved(all_grants_page)

    # Create a page with a history of all lazy consensus proposals
    proposals_page = wb.create_sheet(title="Proposals")
    await write_lazy_consensus_history(proposals_page)

    # Create a page with all free funding transactions
    free_funding_history_page = wb.create_sheet(title="Tips Transactions")
    await write_free_funding_transactions(free_funding_history_page)

    # Save the Excel workbook to a temporary file
    temp_file = io.BytesIO()
    wb.save(temp_file)
    temp_file.seek(0)

    return temp_file, EXPORT_DATA_FILENAME


@client.command(name=EXPORT_COMMAND_NAME)
async def export_command(ctx):
    try:
        # Reply to a non-authorized user
        if not await validate_roles(ctx.message.author):
            # Adding greetings and "cancelled" reactions
            await ctx.message.add_reaction(REACTION_ON_BOT_MENTION)
            # Sending response in DM
            await ctx.message.reply(HELP_MESSAGE_NON_AUTHORIZED_USER)
            return
        # Adding greetings reaction so to show that the command is being processed (it may take a couple of seconds waiting for the user)
        await ctx.message.add_reaction(REACTION_ON_BOT_MENTION)

        # Create the document
        document, filename = await export_xlsx()
        # Send the document to user
        await ctx.message.reply(
            EXPORT_CHANNEL_REPLY,
            file=discord.File(document, filename=filename),
        )

    except Exception as e:
        try:
            # Try replying in Discord
            await ctx.message.reply(
                f"An unexpected error occurred when exporting analytical data. cc {RESPONSIBLE_MENTION}"
            )
        except Exception as e:
            logger.critical("Unable to reply in the chat that a critical error has occurred.")

        logger.critical(
            "Unexpected error in %s while exporting analytical data, channel=%s, message=%s, user=%s",
            __name__,
            ctx.message.channel.id if ctx.message.channel else None,
            ctx.message.id,
            ctx.message.author.mention,
            exc_info=True,
        )
